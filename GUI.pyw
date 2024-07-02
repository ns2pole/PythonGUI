import pandas as pd
from openpyxl import load_workbook

# エクセルファイルの読み込み
file_path = 'employee_test.xlsx'
workbook = load_workbook(file_path)

# シートの取得
sheet = workbook['test']
sheet['D5'] = '新しい値'
# sheet.merge_cells('A3:C3')
workbook.save(file_path)