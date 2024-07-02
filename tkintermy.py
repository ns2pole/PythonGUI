import tkinter as tk
import sqlite3
from openpyxl import load_workbook

class MainWindow(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.pack()

        self.label = tk.Label(self, text="ぼたんを押してね")
        self.label.pack(pady=10)

        self.button = tk.Button(self, text="Say Hello", command=self.rundb)
        self.button.pack()

    def say_hello(self):
        self.label.config(text="Hello!")

    def rundb(self):
        conn = sqlite3.connect("Cosmos.db", isolation_level=None)
        cursor = conn.cursor()

        # usersテーブルを削除して初期化
        cursor.execute("DROP TABLE users")
        # usersテーブルを作成するSQL文
        create_table_query = "CREATE TABLE users (id INTEGER PRIMARY KEY,name TEXT NOT NULL,age INTEGER,email TEXT)"

        # SQLクエリを実行
        cursor.execute(create_table_query)

        # データを挿入するSQL文
        insert_data_queries = [
            "INSERT INTO users (name, age, email) VALUES ('Alice', 30, 'alice@example.com')",
            "INSERT INTO users (name, age, email) VALUES ('Bob', 25, 'bob@example.com')",
            "INSERT INTO users (name, age, email) VALUES ('Charlie', 35, 'charlie@example.com')"
        ]

        # 各SQLクエリを実行してデータを挿入
        for query in insert_data_queries:
            cursor.execute(query)
            
        # データを取得
        select_data_query = "SELECT * FROM users"
        cursor.execute(select_data_query)
        rows = cursor.fetchall()

        # 結果を表示
        for row in rows:
            print(row)


        file_path = 'employee_test.xlsx'
        workbook = load_workbook(file_path)

        # シートの取得
        sheet = workbook['test']
        sheet['D5'] = "rowstest"
        # sheet.merge_cells('A3:C3')
        workbook.save(file_path)

# Tkinter Application Setup
if __name__ == "__main__":
    application = tk.Tk()
    application.title('Hello App')

    window = MainWindow(application)

    application.mainloop()